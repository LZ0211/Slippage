# coding=utf-8
import os,re
from uuid import uuid4 as uuid
from tempfile import mktemp
from chardet import detect
from xlrd import open_workbook
from xlwt import Workbook as xls_Workbook
from openpyxl import load_workbook
from openpyxl import Workbook as xlsx_Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import LineChart,Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

class File:
    @staticmethod
    def temp(data,file=None):
        if file == None:
            file = mktemp() + '.txt'
        open(file,'w+').write(data)
        return file

    def __init__(self,filename=None):
        if not filename == None:
            found = re.match(r'^"(.*)"$',filename)
            if found:
                filename = found[1]
            (filepath, tempfilename) = os.path.split(filename)
            (filename, filetype) = os.path.splitext(tempfilename)
            self.filepath = filepath
            self.filename = filename
            self.filetype = filetype.lower()
            self.file = os.path.join(self.filepath,self.filename+self.filetype)
            self.data = []

    def detectCode(self,path):
	    with open(path, 'rb') as file:
		    data = file.read(200000)
		    dicts = detect(data)
	    return dicts["encoding"]

    def read_data(self):
        if self.filetype == '.txt':
            self.data = self.read_txt_file(self.file)
        elif self.filetype == '.csv':
            self.data = self.read_txt_file(self.file)
        elif self.filetype == '.xls':
            self.data = self.read_xls_file(self.file)
        elif self.filetype == '.xlsx':
            self.data = self.read_xls_file(self.file)
        else:
            raise Exception(self.file + ' is invalid data file!')
        #过滤空白行
        data = []
        for line in self.data:
            if not (None in line):
                data.append(line)
        self.data = data
        return data

    def write_data(self,data):
        if self.filetype == '.txt':
            self.write_txt_file(self.file,data)
        elif self.filetype == '.csv':
            self.write_csv_file(self.file,data)
        elif self.filetype == '.xls':
            self.write_xls_file(self.file,data)
        elif self.filetype == '.xlsx':
            self.write_xlsx_file(self.file,data)

    def save_as(self,file):
        (filepath, tempfilename) = os.path.split(file)
        (filename, filetype) = os.path.splitext(tempfilename)
        filetype = filetype.lower()
        if filetype == '.txt':
            self.write_txt_file(file,self.data)
        elif filetype == '.csv':
            self.write_csv_file(file,self.data)
        elif filetype == '.xls':
            self.write_xls_file(file,self.data)
        elif filetype == '.xlsx':
            self.write_xlsx_file(file,self.data)

    def read_txt_file(self,filename):
        encoding = self.detectCode(filename)
        lines = open(filename,encoding=encoding).read().strip().splitlines()
        lines = filter(lambda line: not re.match(r'[^\s\d\.\+\-]',line),lines)
        points = filter(lambda x:len(x)==2,map(lambda line:list(map(lambda x:float(x),filter(lambda x: not x == '',re.split(r'[,\s]+',line.strip())))),lines))
        return list(points)

    def read_xls_file(self,filename):
        wb = open_workbook(filename)
        ws = wb.sheet_by_index(0)
        points = []
        for i in range(1,ws.nrows):
            line = []
            for j in range(ws.ncols):
                val = ws.cell_value(i, j)
                line.append(val)
            points.append(line)
        return points

    def read_xlsx_file(self,filename):
        wb = load_workbook(filename)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        points = []
        for i in range(1,ws.max_row):
            line = []
            for j in range(ws.max_column):
                val = ws.cell(i+1,j+1).value
                line.append(val)
            points.append(line)
        return points

    def write_txt_file(self,filename,data):
        open(filename,'w+').write('\n'.join(map(lambda x:'\t'.join(map(str,x)),data)))

    def write_csv_file(self,filename,data):
        open(filename,'w+').write('\n'.join(map(lambda x:','.join(map(str,x)),data)))

    def write_xls_file(self,filename,data):
        wb = xls_Workbook(encoding="utf-8")
        ws = wb.add_sheet('Sheet1', cell_overwrite_ok=True)
        for i in range(len(data)):
            line = data[i]
            for j in range(len(line)):
                ws.write(i,j,line[j])
        wb.save(filename)

    def write_xlsx_file(self,filename,data):
        wb = xlsx_Workbook()
        ws = wb.active
        for i in range(len(data)):
            line = data[i]
            for j in range(len(line)):
                ws.cell(i+1, j+1, line[j])
        wb.save(filename)


class Table:
    def __init__(self,stream=None,filename=None):
        self.filename = filename
        if filename == None:
            self.filename = mktemp() + '.xlsx'
        if stream == None:
            self.workbook =  xlsx_Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = '汇总'
            #self.worksheet.protection.password = str(uuid())
            self.rows = 1
            self.init_tab_header()
        else:
            open(self.filename,'wb+').write(stream)
            self.workbook = load_workbook(self.filename)
            self.worksheet = self.workbook.get_sheet_by_name(self.workbook.sheetnames[0])
            #self.worksheet.protection.password = str(uuid())
            self.rows = self.worksheet.max_row + 1


    def init_tab_header(self):
        border = Border(
            left=Side(style='thin',color='FF000000'),
            right=Side(style='thin',color='FF000000'),
            top=Side(style='thin',color='FF000000'),
            bottom=Side(style='thin',color='FF000000')
        )
        fill = PatternFill(
            fill_type='solid',
            fgColor='FFFFFF00'
        )
        font = Font(name='等线',size=14,bold=True)
        self.worksheet.row_dimensions[1].height = 18
        self.worksheet.column_dimensions['A'].width = 9
        self.worksheet.column_dimensions['B'].width = 12
        self.worksheet.column_dimensions['C'].width = 12
        self.worksheet.column_dimensions['D'].width = 9
        self.worksheet.column_dimensions['E'].width = 9
        self.worksheet.column_dimensions['F'].width = 9
        self.worksheet.column_dimensions['G'].width = 9
        self.worksheet.column_dimensions['H'].width = 9
        self.worksheet.column_dimensions['I'].width = 9
        self.worksheet.column_dimensions['J'].width = 9
        self.worksheet.column_dimensions['K'].width = 9
        self.worksheet.column_dimensions['L'].width = 9
        self.worksheet['B1'] = 'Qfull'
        self.worksheet['B1'].border = border
        self.worksheet['B1'].font = font
        font = Font(name='等线',size=11,bold=True)
        self.worksheet['C1'] = 'No.'
        self.worksheet['C1'].border = border
        self.worksheet['C1'].font = font
        font = Font(name='等线',size=11)
        self.worksheet['D1'] = 'Mp'
        self.worksheet['D1'].border = border
        self.worksheet['D1'].fill = fill
        self.worksheet['D1'].font = font
        self.worksheet['E1'] = 'Sp'
        self.worksheet['E1'].border = border
        self.worksheet['E1'].fill = fill
        self.worksheet['E1'].font = font
        self.worksheet['F1'] = 'Mn'
        self.worksheet['F1'].border = border
        self.worksheet['F1'].fill = fill
        self.worksheet['F1'].font = font
        self.worksheet['G1'] = 'Sn'
        self.worksheet['G1'].border = border
        self.worksheet['G1'].fill = fill
        self.worksheet['G1'].font = font
        self.worksheet['H1'] = 'RMSD'
        self.worksheet['H1'].border = border
        self.worksheet['H1'].fill = fill
        self.worksheet['H1'].font = font
        fill = PatternFill(
            fill_type='solid',
            fgColor='FF00B0F0'
        )
        self.worksheet['I1'] = 'LLI'
        self.worksheet['I1'].border = border
        self.worksheet['I1'].fill = fill
        self.worksheet['I1'].font = font
        self.worksheet['J1'] = 'LAM-PE'
        self.worksheet['J1'].border = border
        self.worksheet['J1'].fill = fill
        self.worksheet['J1'].font = font
        self.worksheet['K1'] = 'LAM-NE'
        self.worksheet['K1'].border = border
        self.worksheet['K1'].fill = fill
        self.worksheet['K1'].font = font
        self.worksheet['L1'] = 'Total Loss'
        self.worksheet['L1'].border = border
        self.worksheet['L1'].fill = fill
        self.worksheet['L1'].font = font
        self.rows += 1
    
    def write_params(self,data):
        border = Border(
            left=Side(style='thin',color='FF000000'),
            right=Side(style='thin',color='FF000000'),
            top=Side(style='thin',color='FF000000'),
            bottom=Side(style='thin',color='FF000000')
        )
        font = Font(name='等线',size=11)
        cells = 'B,C,D,E,F,G,H'.split(',')
        idx = 0
        for cell in cells:
            cell += str(self.rows)
            self.worksheet[cell] = data[idx]
            self.worksheet[cell].border = border
            self.worksheet[cell].font = font
            idx += 1
        cell_id = 'I%s' % self.rows
        self.worksheet[cell_id] = '=(($E$2-$G$2+$F$2*355.8)-(E%s-G%s+F%s*355.8))/($E$2-$G$2+$F$2*355.8)' % (self.rows,self.rows,self.rows)
        self.worksheet[cell_id].border = border
        self.worksheet[cell_id].font = font
        self.worksheet[cell_id].number_format = '0.00%'
        cell_id = 'J%s' % self.rows
        self.worksheet[cell_id] = '=($D$2-D%s)/$D$2' % self.rows
        self.worksheet[cell_id].border = border
        self.worksheet[cell_id].font = font
        self.worksheet[cell_id].number_format = '0.00%'
        cell_id = 'K%s' % self.rows
        self.worksheet[cell_id] = '=($F$2-F%s)/$F$2' % self.rows
        self.worksheet[cell_id].border = border
        self.worksheet[cell_id].font = font
        self.worksheet[cell_id].number_format = '0.00%'
        cell_id = 'L%s' % self.rows
        self.worksheet[cell_id] = '=($B$2-B%s)/$B$2' % self.rows
        self.worksheet[cell_id].border = border
        self.worksheet[cell_id].font = font
        self.worksheet[cell_id].number_format = '0.00%'
        self.rows += 1
    
    def insert_graph(self):
        c1 = LineChart()
        c1.title = "Degradation Mode"
        #c1.style = 12
        c1.y_axis.title = 'degradation percent'
        c1.x_axis.title = ''
        data = Reference(self.worksheet, min_col=9, min_row=1, max_col=12, max_row=self.rows-1)
        cats = Reference(self.worksheet, min_col=3, min_row=2, max_row=self.rows-1)
        c1.add_data(data, titles_from_data=True)
        c1.set_categories(cats)
        for serie in c1.series:
            serie.graphicalProperties.line.width = 24000
        sgp = GraphicalProperties(ln=LineProperties(noFill=True))
        c1.y_axis.majorGridlines.spPr = sgp
        self.worksheet.add_chart(c1, "C%s"%self.rows)

    def view_file(self,temp=None):
        if temp == None:
            temp = self.filename
        self.save_file(temp)
        new_tab = Table(open(temp,'rb').read())
        new_tab.insert_graph()
        new_tab.workbook.security.workbookPassword = str(uuid())
        new_tab.workbook.security.lockStructure = True
        new_tab.worksheet.protection.password = str(uuid())
        new_tab.worksheet.protect = True
        new_tab.worksheet.protection.enable()
        new_tab.save_file(temp)
        return temp
        #subprocess.Popen(['%s' % temp],shell=True)

    #原始数据，不带图表
    def save_file(self,filename=None):
        if filename == None:
            filename = self.filename
        self.workbook.save(filename)

    def export_file(self,filename=None):
        if filename == None:
            filename = self.filename
        self.save_file()
        new_tab = Table(open(self.filename,'rb').read())
        new_tab.insert_graph()
        new_tab.save_file(filename)

